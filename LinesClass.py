import openpyxl as op
import random
import pygame as pg


class DrawLines:
    def __init__(self, tempetarute, steps):
        self.lines_coord = []
        self.lines_coord_2 = []
        self.lines_coord_3 = []
        self.graf_space = pg.Rect(200, 50, 1000, 620)
        self.prev_step = steps[0]
        self.prev_pos = (200, 670)
        self.temp = tempetarute
        self.count = 1
        self.line_steps = steps
        self.count_of_saves = 1

    def save(self):
        if self.lines_coord == []:
            pass
        else:
            if self.count == 1:
                wb = op.Workbook()
                sheet = wb.active
                i = 0
                for y in self.lines_coord:
                    i += 1
                    sheet.cell(row=i+1, column=1).value = i
                    sheet.cell(row=i+1, column=self.count + 1).value = (670 - y[1]) * (int(self.temp) / 620)
                sheet.cell(row=1, column=1).value = 0
                sheet.cell(row=2, column=1).value = 1
                sheet.cell(row=1, column=self.count + 1).value = random.uniform(18, 20)
                self.count += 1
                wb.save("Test" + str(self.count_of_saves) + ".xlsx")
            else:
                wb = op.load_workbook("Test" + str(self.count_of_saves) + ".xlsx")
                sheet = wb.active
                i = 0
                for y in self.lines_coord:
                    i += 1
                    sheet.cell(row=i+1, column=1).value = i
                    sheet.cell(row=i+1, column=self.count + 1).value = (670 - y[1]) * (int(self.temp) / 620)
                sheet.cell(row=1, column=1).value = 0
                sheet.cell(row=2, column=1).value = 1
                sheet.cell(row=1, column=self.count + 1).value = random.uniform(18, 20)
                self.count += 1
                wb.save("Test" + str(self.count_of_saves) + ".xlsx")
